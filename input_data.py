from typing import List, Optional, Dict, Union

import openpyxl

import pandas as pd


def extract_links(
    excel_file: str,
    url_col_header: str = "ProductURL",
    unique_id_header: str = "ItemID",
    sheet_name: Optional[str] = None,
) -> pd.DataFrame:

    raw_book = openpyxl.load_workbook(excel_file)

    if sheet_name:
        raw_sheet = raw_book[sheet_name]
    else:
        # If sheet name is not specified use the first sheet in the workbook
        raw_sheet = raw_book[raw_book.sheetnames[0]]

    # Build the header dictionary
    top_row: tuple = next(raw_sheet.rows)
    headers: Dict[str, int] = {}
    for cell in top_row:
        headers[cell.value] = cell.col_idx
        if cell.value == url_col_header:
            url_col_letter = cell.column_letter

    # Extract the urls
    url_column = raw_sheet[url_col_letter]

    url_list: List[Dict[str, Union[int, str]]] = []

    # We have to skip the first row in the column as this is the headers but Excel indexes
    # from 1 (not 0 like a proper program) so we have to start the row count from 2.
    for row, cell in enumerate(url_column[1:], start=2):

        # Get the unique identifier value from the sheet
        unique_id = raw_sheet.cell(column=headers[unique_id_header], row=row).value
        if unique_id:
            int(unique_id)
        else:
            unique_id = "NA"

        row_dict: Dict[str, Union[str, int]] = {unique_id_header: unique_id}

        if cell.hyperlink:
            row_dict["ProductURL"] = cell.hyperlink.target
        else:
            row_dict["ProductURL"] = "NA"

        url_list.append(row_dict)

    output: pd.DataFrame = pd.DataFrame(url_list)

    # Some of the rows will be empty if the sheet had extra cells so we drop them
    output = output[output[unique_id_header] != "NA"]

    return output


def get_input_dataframe(
    excel_file: str,
    url_col_header: str = "ProductURL",
    unique_id_header: str = "ItemID",
    sheet_name: Optional[str] = None,
    url_col_new_name: str = "ProductDescription",
) -> pd.DataFrame:

    urls: pd.DataFrame = extract_links(
        excel_file=excel_file,
        url_col_header=url_col_header,
        unique_id_header=unique_id_header,
        sheet_name=sheet_name,
    )

    raw_data: pd.DataFrame = pd.read_excel(excel_file)
    raw_data = raw_data.rename(columns={url_col_header: url_col_new_name})

    merged = raw_data.merge(urls, on=unique_id_header)

    return merged
