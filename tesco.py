import re
import warnings

from typing import List, Optional, Dict, Union, Pattern, Match

import requests
import openpyxl

import pandas as pd

PRODUCT_CODE_PATTERN: Pattern = re.compile(r"\d{9}")

PRODUCT_SEARCH_URL: str = "https://dev.tescolabs.com/grocery/products"

PRODUCT_DATA_URL: str = "https://dev.tescolabs.com/product"


def product_search(api_key: str, query: str, limit: int = 10, offset: int = 0):

    headers = {"Ocp-Apim-Subscription-Key": api_key}

    params = {"query": query, "limit": limit, "offset": offset}

    response: requests.Response = requests.get(
        PRODUCT_SEARCH_URL, headers=headers, params=params
    )

    response.raise_for_status()

    return response.json()


def product_data(api_key: str, product_id: str):

    headers = {"Ocp-Apim-Subscription-Key": api_key}

    params = {"tpnc": product_id}

    response: requests.Response = requests.get(
        PRODUCT_DATA_URL, headers=headers, params=params
    )

    response.raise_for_status()

    raw = response.json()["products"]

    if len(raw) > 1:
        warnings.warn(f"More than one product returned for item {product_id}")

    return raw[0]


def extract_links(
    excel_file: str,
    hyperlink_col_header: str = "ProductURL",
    new_url_col_header: str = "ProductURL",
    unique_id_header: str = "ItemID",
    sheet_name: Optional[str] = None,
) -> pd.DataFrame:

    raw_book = openpyxl.load_workbook(excel_file)

    if sheet_name:
        raw_sheet = raw_book[sheet_name]
    else:
        # If sheet name is not specified use the first sheet in the workbook
        raw_sheet = raw_book[raw_book.sheetnames[0]]

    # Build the header dictionary
    top_row: tuple = next(raw_sheet.rows)
    headers: Dict[str, int] = {}
    for cell in top_row:
        headers[cell.value] = cell.col_idx
        if cell.value == hyperlink_col_header:
            url_col_letter = cell.column_letter

    # Extract the urls
    url_column = raw_sheet[url_col_letter]

    url_list: List[Dict[str, Union[int, str]]] = []

    # We have to skip the first row in the column as this is the headers but Excel indexes
    # from 1 (not 0 like a proper program) so we have to start the row count from 2.
    for row, cell in enumerate(url_column[1:], start=2):

        # Get the unique identifier value from the sheet
        unique_id = raw_sheet.cell(column=headers[unique_id_header], row=row).value
        if unique_id:
            int(unique_id)
        else:
            unique_id = "NA"

        row_dict: Dict[str, Union[str, int]] = {unique_id_header: unique_id}

        if cell.hyperlink:
            row_dict[new_url_col_header] = cell.hyperlink.target
        else:
            row_dict[new_url_col_header] = "NA"

        url_list.append(row_dict)

    output: pd.DataFrame = pd.DataFrame(url_list)

    # Some of the rows will be empty if the sheet had extra cells so we drop them
    output = output[~output[unique_id_header].isna()]

    return output


def extract_product_code(url: str) -> str:

    code_match: Optional[Match] = re.search(PRODUCT_CODE_PATTERN, url)

    if not code_match:
        return "NA"

    return code_match.group()


def extract_nutrition(product_dict: dict) -> pd.DataFrame:

    return pd.DataFrame(product_dict["calcNutrition"]["calcNutrients"])


def get_input_dataframe(
    excel_file: str,
    hyperlink_col_header: str = "ProductURL",
    hyperlink_col_new_name: str = "ProductDescription",
    new_url_col_header: str = "ProductURL",
    new_product_code_header: str = "ProductCode",
    unique_id_header: str = "ItemID",
    sheet_name: Optional[str] = None,
) -> pd.DataFrame:

    urls: pd.DataFrame = extract_links(
        excel_file=excel_file,
        hyperlink_col_header=hyperlink_col_header,
        new_url_col_header=new_url_col_header,
        unique_id_header=unique_id_header,
        sheet_name=sheet_name,
    )

    # Extract the product code from the url and add it as a column in the dataframe
    urls[new_product_code_header] = urls[hyperlink_col_header].apply(
        extract_product_code
    )

    raw_data: pd.DataFrame = pd.read_excel(excel_file)
    raw_data = raw_data.rename(columns={hyperlink_col_header: hyperlink_col_new_name})

    merged = raw_data.merge(urls, on=unique_id_header)

    return merged
